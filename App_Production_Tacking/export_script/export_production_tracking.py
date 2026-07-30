import json
import os
import sys
import logging
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook


def carica_config(path="config.json"):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def setup_logging(config):
    cfg = config["export"]
    if cfg.get("log_enabled", False):
        log_path = Path(cfg.get("log_path", "logs"))
        log_path.mkdir(parents=True, exist_ok=True)
        log_file = log_path / f"export_{datetime.now().strftime('%Y%m%d')}.log"
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] %(message)s",
            handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler()],
        )
    else:
        logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    logging.info("=== Export Production Tracking avviato ===")


def fetch_view(url, anon_key, view_name, max_righe=100000):
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    full_url = f"{url}/rest/v1/{view_name}"
    params = {"limit": max_righe}

    try:
        resp = requests.get(full_url, headers=headers, params=params, timeout=120)
        resp.raise_for_status()
        data = resp.json()
        logging.info(f"  {view_name}: {len(data)} righe lette")
        return data
    except requests.exceptions.RequestException as e:
        logging.error(f"  {view_name}: errore HTTP -> {e}")
        if hasattr(e, "response") and e.response is not None:
            logging.error(f"    Risposta: {e.response.text[:500]}")
        return []


def scrivi_xlsx(config, dati):
    cfg = config["export"]
    output_dir = Path(cfg["output_path"])
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"{cfg.get('filename_prefix', 'Export_Production_Tracking_')}{timestamp}.xlsx"
    filepath = output_dir / filename

    wb = Workbook()
    wb.remove(wb.active)

    mappatura_view = {
        "order_history_csv_view": "Storico_Ordini",
        "bem_job_view": "BEM_Job",
        "bem_forno_view": "BEM_Forno",
    }

    for view_name, sheet_name in mappatura_view.items():
        righe = dati.get(view_name, [])
        ws = wb.create_sheet(title=sheet_name)

        if not righe:
            logging.warning(f"  Nessun dato per {view_name}, foglio vuoto")
            continue

        colonne = list(righe[0].keys())
        ws.append(colonne)

        for riga in righe:
            ws.append([riga.get(col, "") for col in colonne])

        for i in range(1, len(colonne) + 1):
            cell = ws.cell(row=1, column=i)
            max_len = len(str(cell.value))
            for row in ws.iter_rows(min_row=2, min_col=i, max_col=i, values_only=True):
                for val in row:
                    if val:
                        max_len = max(max_len, len(str(val)))
            ws.column_dimensions[chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)].width = min(max_len + 3, 55)

    wb.save(filepath)
    logging.info(f"File salvato: {filepath}")
    return filepath


def aggiorna_aziendale(config, filepath_export):
    cfg = config.get("aziendale", {})
    if not cfg.get("enabled", False):
        logging.info("Aggiornamento DB aziendale disabilitato, skippo.")
        return

    xls_path = cfg.get("xls_path", "")
    if not xls_path or not os.path.exists(xls_path):
        logging.warning(f"File aziendale non trovato: {xls_path}")
        return

    col_ordine = cfg.get("colonna_ordine", "OdL")
    col_stato = cfg.get("colonna_stato", "Ultimo_avanzamento")
    col_data = cfg.get("colonna_data", "Data_ultimo_avanzamento")
    col_reparto = cfg.get("colonna_reparto", "Reparto_corrente")

    logging.info(f"Aggiorno DB aziendale: {xls_path}")

    wb_export = load_workbook(filepath_export)
    ws_storico = wb_export["Storico_Ordini"]

    ultimi_avanzamenti = {}
    header_storico = [str(c.value).strip() for c in ws_storico[1]]
    idx_odl = header_storico.index("OdL") if "OdL" in header_storico else 2
    idx_data_op = header_storico.index("Data_operazione") if "Data_operazione" in header_storico else 0
    idx_tipo_op = header_storico.index("Tipo_operazione") if "Tipo_operazione" in header_storico else 5
    idx_al_reparto = header_storico.index("Al_reparto") if "Al_reparto" in header_storico else 7

    for row in ws_storico.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx_odl]:
            continue
        odl = str(row[idx_odl]).strip()
        data_op = str(row[idx_data_op]).strip() if row[idx_data_op] else ""
        operazione = str(row[idx_tipo_op]).strip() if row[idx_tipo_op] else ""
        al_reparto = str(row[idx_al_reparto]).strip() if row[idx_al_reparto] else ""

        if odl not in ultimi_avanzamenti or data_op > ultimi_avanzamenti[odl].get("data", ""):
            ultimi_avanzamenti[odl] = {"data": data_op, "operazione": operazione, "reparto": al_reparto}

    wb_az = load_workbook(xls_path)
    ws_az = wb_az.active

    header_az = [str(c.value).strip() if c.value else "" for c in ws_az[1]]
    logging.info(f"Colonne file aziendale: {header_az}")

    try:
        i_ordine = header_az.index(col_ordine)
    except ValueError:
        logging.error(f"Colonna '{col_ordine}' non trovata nel file aziendale.")
        return
    i_stato = header_az.index(col_stato) if col_stato in header_az else None
    i_data = header_az.index(col_data) if col_data in header_az else None
    i_reparto = header_az.index(col_reparto) if col_reparto in header_az else None

    aggiornati = 0
    for row in ws_az.iter_rows(min_row=2, max_row=ws_az.max_row):
        cell = row[i_ordine]
        odl = str(cell.value).strip() if cell.value else ""
        if odl in ultimi_avanzamenti:
            info = ultimi_avanzamenti[odl]
            if i_stato is not None:
                row[i_stato].value = info["operazione"]
            if i_data is not None:
                row[i_data].value = info["data"]
            if i_reparto is not None:
                row[i_reparto].value = info["reparto"]
            aggiornati += 1

    wb_az.save(xls_path)
    logging.info(f"DB aziendale aggiornato: {aggiornati} righe modificate")


def main():
    config = carica_config()
    setup_logging(config)
    supabase = config["supabase"]

    views = ["order_history_csv_view", "bem_job_view", "bem_forno_view"]
    dati = {}

    for view in views:
        logging.info(f"Fetching {view}...")
        dati[view] = fetch_view(supabase["url"], supabase["anon_key"], view)

    if not any(dati.values()):
        logging.error("Nessun dato ricevuto da Supabase. Verifica credenziali e connettivita.")
        sys.exit(1)

    filepath = scrivi_xlsx(config, dati)
    aggiorna_aziendale(config, filepath)

    logging.info("=== Export completato ===")


if __name__ == "__main__":
    main()
